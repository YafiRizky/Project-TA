import openpyxl
import os

folder = r'c:\laragon\www\TA\Data Model\Data Mentah'
output = r'c:\laragon\www\TA\Data Model\semua_data_extracted.txt'

with open(output, 'w', encoding='utf-8') as f:
    for fname in sorted(os.listdir(folder)):
        if not fname.endswith('.xlsx'):
            continue
        filepath = os.path.join(folder, fname)
        wb = openpyxl.load_workbook(filepath)
        f.write(f'\n{"="*70}\n')
        f.write(f'FILE: {fname}\n')
        f.write(f'{"="*70}\n')
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f'\n--- Sheet: {sheet_name} ---\n')
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(c) if c is not None else '' for c in row]
                if all(c == '' for c in cells):
                    continue
                f.write(f'  R{i}: {cells}\n')
        wb.close()

print(f'Done! Written to {output}')
