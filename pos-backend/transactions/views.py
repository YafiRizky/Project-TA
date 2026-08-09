"""
Transaction views for POS system.
Handles checkout, transaction management, void/refund, and summaries.
"""
import logging
from rest_framework import viewsets, filters, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from django.utils import timezone
from django.db.models import Sum, Count, Q, F
from django.db import transaction as db_transaction, IntegrityError
from decimal import Decimal, InvalidOperation
from uuid import uuid4
from .models import Transaction, TransactionItem
from .serializers import TransactionSerializer, TransactionCreateSerializer, TransactionItemSerializer
from products.models import Product
from inventory.models import ProductBatch, InventoryMovement
from accounts.permissions import IsBusinessAdmin
from auditlog.utils import log_action
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

logger = logging.getLogger(__name__)

# Configurable limits
MAX_DISCOUNT_PERCENT = Decimal('50')  # Max 50% diskon per transaksi


class TransactionPagination(PageNumberPagination):
    """Pagination untuk list transaksi. Default 20 per page, max 5000."""
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 5000


class TransactionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Transaction (Sales) CRUD and POS operations
    Automatically filters by user's business
    Supports cart checkout and transaction management
    
    Actions:
    - list: GET /api/transactions/transactions/
    - create: POST /api/transactions/transactions/ (full transaction)
    - retrieve: GET /api/transactions/transactions/{id}/
    - update: PUT /api/transactions/transactions/{id}/
    - destroy: DELETE /api/transactions/transactions/{id}/
    - checkout: POST /api/transactions/transactions/checkout/ (POS cart)
    - void: POST /api/transactions/transactions/{id}/void/ (admin only)
    - daily_summary: GET /api/transactions/transactions/daily_summary/
    - payment_summary: GET /api/transactions/transactions/payment_summary/
    """
    permission_classes = [IsAuthenticated]
    pagination_class = TransactionPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['transaction_code', 'cashier_name']
    ordering_fields = ['transaction_date', 'total_amount', 'status']
    ordering = ['-transaction_date']
    
    def get_permissions(self):
        """Enforce admin-only for destructive actions"""
        if self.action in ['update', 'partial_update', 'destroy']:
            return [IsAuthenticated(), IsBusinessAdmin()]
        return [IsAuthenticated()]
    
    def get_serializer_class(self):
        """Use different serializer for checkout action"""
        if self.action == 'checkout':
            return TransactionCreateSerializer
        return TransactionSerializer
    
    def get_queryset(self):
        """
        Filter transactions by authenticated user's business
        """
        if hasattr(self.request.user, 'business'):
            queryset = Transaction.objects.filter(
                business=self.request.user.business
            ).prefetch_related('items', 'items__product', 'items__batch')
            
            # Filter by status if provided
            status_filter = self.request.query_params.get('status')
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            # Filter by payment method if provided
            payment_method = self.request.query_params.get('payment_method')
            if payment_method:
                queryset = queryset.filter(payment_method=payment_method)
            
            # Filter by date range if provided
            start_date = self.request.query_params.get('start_date')
            end_date = self.request.query_params.get('end_date')
            if start_date:
                queryset = queryset.filter(transaction_date__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(transaction_date__date__lte=end_date)
            
            return queryset
        return Transaction.objects.none()
    
    def _generate_transaction_code(self, business, attempt=0):
        """Generate sequential, collision-free transaction code.
        Format: TRX-YYMMDD-NNNNN (sequential per day, GLOBAL across all businesses)
        Uses retry logic to handle race conditions.
        """
        import uuid
        today = timezone.now().strftime('%y%m%d')
        prefix = f"TRX-{today}-"
        
        # Query ALL transactions with this prefix (global, not per-business)
        # to prevent duplicate key across different businesses
        last_trx = Transaction.objects.filter(
            transaction_code__startswith=prefix
        ).order_by('-id').first()  # order by -id is more reliable than -transaction_code
        
        next_num = 1
        if last_trx:
            try:
                last_num = int(last_trx.transaction_code.split('-')[-1])
                next_num = last_num + 1 + attempt  # Add attempt offset for retries
            except (ValueError, IndexError):
                next_num = 1 + attempt
        else:
            next_num = 1 + attempt
        
        # Safety: if too many retries, add random hex to guarantee uniqueness
        if attempt >= 5:
            return f"{prefix}{next_num:05d}-{uuid.uuid4().hex[:4].upper()}"
        
        return f"{prefix}{next_num:05d}"
    
    def perform_create(self, serializer):
        """Auto-assign business and generate transaction code"""
        transaction_code = self._generate_transaction_code(self.request.user.business)
        serializer.save(
            business=self.request.user.business,
            transaction_code=transaction_code
        )
    
    def update(self, request, *args, **kwargs):
        """Block editing completed/voided transactions"""
        instance = self.get_object()
        if instance.status in ('COMPLETED', 'VOIDED'):
            return Response({'error': 'Transaksi yang sudah selesai/void tidak dapat diubah.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Block deleting completed/voided transactions"""
        instance = self.get_object()
        if instance.status in ('COMPLETED', 'VOIDED'):
            return Response({'error': 'Transaksi yang sudah selesai/void tidak dapat dihapus.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)
    
    @action(detail=False, methods=['post'])
    def checkout(self, request):
        """
        POS Checkout - Process cart items into transaction
        Expects: {items: [{product_id, quantity, price_per_unit, discount?}], payment_method, amount_paid, ...}
        
        Security features:
        - Idempotency key to prevent double submissions
        - Price validation against database
        - Discount limits (max 50%)
        - Atomic stock locking (no race conditions)
        """
        if not hasattr(request.user, 'business'):
            return Response({'error': 'Tidak ada bisnis yang terkait'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = TransactionCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        business = request.user.business
        items_data = serializer.validated_data['items']
        
        # ===== IDEMPOTENCY CHECK (B4) =====
        idempotency_key = request.data.get('idempotency_key') or request.headers.get('X-Idempotency-Key')
        if idempotency_key:
            existing = Transaction.objects.filter(idempotency_key=idempotency_key).first()
            if existing:
                logger.info('Duplicate checkout prevented (key: %s)', idempotency_key)
                result_serializer = TransactionSerializer(existing)
                return Response(result_serializer.data, status=status.HTTP_200_OK)
        
        # Build expiry filter to exclude expired batches
        expiry_filter = Q(expiry_date__isnull=True) | Q(expiry_date__gt=timezone.localdate())
        
        # ===== ATOMIC TRANSACTION WITH FULL LOCKING =====
        try:
            with db_transaction.atomic():
                total_before_discount = Decimal('0.00')
                validated_items = []
                
                for item in items_data:
                    # Validate product exists and belongs to this business
                    try:
                        product = Product.objects.get(id=item['product_id'], business=business)
                    except Product.DoesNotExist:
                        return Response({
                            'error': f'Produk dengan ID {item["product_id"]} tidak ditemukan.'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # ===== PRICE VALIDATION (B2) =====
                    submitted_price = Decimal(str(item['price_per_unit']))
                    if submitted_price != product.selling_price:
                        logger.warning(
                            'Price mismatch for %s: submitted=%s, database=%s (user: %s)',
                            product.name, submitted_price, product.selling_price, request.user.username
                        )
                        return Response({
                            'error': f'Harga {product.name} tidak sesuai. Harga saat ini: Rp {product.selling_price:,.0f}. Silakan refresh halaman.',
                            'product_id': product.id,
                            'expected_price': str(product.selling_price),
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # ===== DISCOUNT VALIDATION (B3) =====
                    item_discount = Decimal(str(item.get('discount', 0)))
                    if item_discount < 0:
                        return Response({'error': 'Diskon tidak boleh negatif.'}, status=status.HTTP_400_BAD_REQUEST)
                    
                    item_subtotal = Decimal(str(item['quantity'])) * submitted_price
                    max_item_discount = item_subtotal * MAX_DISCOUNT_PERCENT / 100
                    if item_discount > max_item_discount:
                        return Response({
                            'error': f'Diskon {product.name} melebihi batas ({MAX_DISCOUNT_PERCENT}%). Maksimum: Rp {max_item_discount:,.0f}'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    total_before_discount += item_subtotal
                    
                    # ===== STOCK CHECK WITH LOCKING (B1) =====
                    # Lock batches FIRST, then check availability (prevents race condition)
                    locked_batches = list(
                        ProductBatch.objects.select_for_update().filter(
                            product=product,
                            business=business,
                            status='ACTIVE'
                        ).filter(expiry_filter).order_by('purchase_date')
                    )
                    
                    available_stock = sum(b.quantity for b in locked_batches)
                    
                    if item['quantity'] > available_stock:
                        if available_stock == 0:
                            return Response({
                                'error': f'Stok {product.name} sudah habis.',
                                'product_id': product.id,
                                'available': 0,
                                'requested': item['quantity']
                            }, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            return Response({
                                'error': f'Stok {product.name} tidak mencukupi. Tersedia: {available_stock}, diminta: {item["quantity"]}',
                                'product_id': product.id,
                                'available': available_stock,
                                'requested': item['quantity']
                            }, status=status.HTTP_400_BAD_REQUEST)
                    
                    validated_items.append({
                        'product': product,
                        'quantity': item['quantity'],
                        'price_per_unit': submitted_price,
                        'discount': item_discount,
                        'locked_batches': locked_batches,
                    })
                
                # ===== TOTAL DISCOUNT VALIDATION (B3) =====
                discount_amount = Decimal(str(serializer.validated_data.get('discount_amount', 0)))
                if discount_amount < 0:
                    return Response({'error': 'Diskon total tidak boleh negatif.'}, status=status.HTTP_400_BAD_REQUEST)
                
                max_total_discount = total_before_discount * MAX_DISCOUNT_PERCENT / 100
                if discount_amount > max_total_discount:
                    return Response({
                        'error': f'Diskon total melebihi batas ({MAX_DISCOUNT_PERCENT}%). Maksimum: Rp {max_total_discount:,.0f}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                total_amount = total_before_discount - discount_amount
                amount_paid = Decimal(str(serializer.validated_data['amount_paid']))
                
                if amount_paid < total_amount:
                    return Response({
                        'error': f'Pembayaran kurang. Total: Rp {total_amount:,.0f}, dibayar: Rp {amount_paid:,.0f}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Generate sequential transaction code (with retry attempt offset)
                checkout_attempt = getattr(request, '_checkout_attempt', 0)
                transaction_code = self._generate_transaction_code(business, attempt=checkout_attempt)
                
                # Create transaction
                trx = Transaction.objects.create(
                    business=business,
                    transaction_code=transaction_code,
                    total_amount=total_amount,
                    discount_amount=discount_amount,
                    payment_method=serializer.validated_data.get('payment_method', 'CASH'),
                    amount_paid=amount_paid,
                    change_amount=amount_paid - total_amount,
                    cashier_name=serializer.validated_data.get('cashier_name', ''),
                    notes=serializer.validated_data.get('notes', ''),
                    status='COMPLETED',
                    idempotency_key=idempotency_key,
                )
                
                # Create transaction items and update inventory (FIFO)
                for vi in validated_items:
                    product = vi['product']
                    quantity_needed = vi['quantity']
                    price_per_unit = vi['price_per_unit']
                    item_discount = vi['discount']
                    
                    # Allocate from locked batches (FIFO) — already locked above
                    for batch in vi['locked_batches']:
                        if quantity_needed <= 0:
                            break
                        
                        quantity_from_batch = min(batch.quantity, quantity_needed)
                        
                        # Create transaction item
                        TransactionItem.objects.create(
                            transaction=trx,
                            product=product,
                            batch=batch,
                            quantity=quantity_from_batch,
                            price_per_unit=price_per_unit,
                            subtotal=Decimal(str(quantity_from_batch)) * price_per_unit,
                            discount=item_discount,
                            cost_per_unit=batch.purchase_cost or Decimal('0.00'),
                        )
                        
                        # Reduce batch quantity and update status
                        batch.quantity -= quantity_from_batch
                        if batch.quantity == 0:
                            batch.status = 'DEPLETED'
                        batch.save()
                        
                        # Log inventory movement
                        InventoryMovement.objects.create(
                            business=business,
                            batch=batch,
                            movement_type='OUT',
                            quantity=quantity_from_batch,
                            notes=f"POS Sale - {product.name}",
                            reference_id=transaction_code
                        )
                        
                        quantity_needed -= quantity_from_batch
                    
                    # Safety net: should never happen after locked pre-validation
                    if quantity_needed > 0:
                        raise Exception(f'Stok {product.name} habis saat proses checkout.')
            
            logger.info('Checkout successful: %s (user: %s, total: %s)', 
                       transaction_code, request.user.username, total_amount)
            
            # Audit log: checkout
            log_action(request, 'CHECKOUT', 'Transaksi',
                       f'Checkout {transaction_code} ({len(items_data)} item, Rp {total_amount:,.0f}, {serializer.validated_data.get("payment_method", "CASH")})',
                       target_id=trx.id,
                       new_data={'transaction_code': transaction_code, 'total': str(total_amount), 'payment': serializer.validated_data.get('payment_method', 'CASH')})
            
            # Return completed transaction
            result_serializer = TransactionSerializer(trx)
            return Response(result_serializer.data, status=status.HTTP_201_CREATED)
            
        except IntegrityError as e:
            err_str = str(e)
            if 'idempotency_key' in err_str:
                # Race condition on idempotency key — return existing transaction
                existing = Transaction.objects.filter(idempotency_key=idempotency_key).first()
                if existing:
                    result_serializer = TransactionSerializer(existing)
                    return Response(result_serializer.data, status=status.HTTP_200_OK)
            if 'transaction_code' in err_str:
                # Race condition on transaction_code — retry with attempt counter
                attempt = getattr(request, '_checkout_attempt', 0) + 1
                if attempt < 5:
                    request._checkout_attempt = attempt
                    logger.warning('Transaction code collision (attempt %d), retrying...', attempt)
                    return self.checkout(request)
            logger.error('Checkout IntegrityError: %s', err_str)
            return Response({'error': 'Terjadi kesalahan database. Silakan coba lagi.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            logger.error('Checkout error: %s', str(e))
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def void(self, request, pk=None):
        """
        Void a completed transaction (admin only)
        - Returns stock to original batches
        - Logs inventory movement as VOID_RETURN
        - Records who voided and why
        """
        # Permission check: admin only
        if not hasattr(request.user, 'role') or request.user.role != 'admin':
            return Response({'error': 'Hanya admin yang dapat membatalkan transaksi.'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            trx = Transaction.objects.get(pk=pk, business=request.user.business)
        except Transaction.DoesNotExist:
            return Response({'error': 'Transaksi tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
        
        if trx.status == 'VOIDED':
            return Response({'error': 'Transaksi ini sudah pernah di-void.'}, status=status.HTTP_400_BAD_REQUEST)
        
        if trx.status != 'COMPLETED':
            return Response({'error': 'Hanya transaksi COMPLETED yang dapat di-void.'}, status=status.HTTP_400_BAD_REQUEST)
        
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response({'error': 'Alasan void wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)
        
        with db_transaction.atomic():
            # Return stock to batches
            for item in trx.items.all():
                if item.batch:
                    batch = ProductBatch.objects.select_for_update().get(pk=item.batch.pk)
                    batch.quantity += item.quantity
                    if batch.status == 'DEPLETED':
                        batch.status = 'ACTIVE'
                    batch.save()
                    
                    # Log inventory return movement
                    InventoryMovement.objects.create(
                        business=trx.business,
                        batch=batch,
                        movement_type='IN',
                        quantity=item.quantity,
                        notes=f"VOID RETURN - {item.product.name} (TRX: {trx.transaction_code})",
                        reference_id=f"VOID-{trx.transaction_code}"
                    )
            
            # Update transaction status
            trx.status = 'VOIDED'
            trx.voided_at = timezone.now()
            trx.voided_by = request.user.full_name or request.user.username
            trx.void_reason = reason
            trx.save()
        
        logger.info('Transaction voided: %s by %s (reason: %s)', 
                    trx.transaction_code, request.user.username, reason)
        
        # Audit log: void
        log_action(request, 'VOID', 'Transaksi',
                   f'Void transaksi {trx.transaction_code} (alasan: {reason})',
                   target_id=trx.id,
                   old_data={'status': 'COMPLETED', 'total': str(trx.total_amount)},
                   new_data={'status': 'VOIDED', 'reason': reason})
        
        result_serializer = TransactionSerializer(trx)
        return Response({
            'message': f'Transaksi {trx.transaction_code} berhasil di-void.',
            'transaction': result_serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def daily_summary(self, request):
        """Get daily sales summary"""
        if not hasattr(request.user, 'business'):
            return Response({'error': 'Tidak ada bisnis yang terkait'}, status=status.HTTP_400_BAD_REQUEST)
        
        today = timezone.localdate()
        transactions = self.get_queryset().filter(
            transaction_date__date=today,
            status='COMPLETED'
        )
        
        # Use ORM aggregation instead of Python-side loops
        aggregated = transactions.aggregate(
            total_revenue=Sum('total_amount'),
            total_items=Sum('items__quantity')
        )
        total_revenue = aggregated['total_revenue'] or Decimal('0.00')
        total_items = aggregated['total_items'] or 0
        total_transactions = transactions.count()
        
        # Use values().annotate() for payment breakdown
        payment_breakdown_qs = transactions.values('payment_method').annotate(
            total=Sum('total_amount')
        )
        payment_breakdown = {
            entry['payment_method']: str(entry['total'])
            for entry in payment_breakdown_qs
        }
        
        return Response({
            'date': str(today),
            'total_revenue': str(total_revenue),
            'transaction_count': total_transactions,
            'item_count': total_items,
            'payment_breakdown': payment_breakdown,
            'average_transaction': str(total_revenue / total_transactions if total_transactions > 0 else 0)
        })
    
    @action(detail=False, methods=['get'])
    def payment_summary(self, request):
        """Get payment method summary"""
        if not hasattr(request.user, 'business'):
            return Response({'error': 'Tidak ada bisnis yang terkait'}, status=status.HTTP_400_BAD_REQUEST)
        
        transactions = self.get_queryset().filter(status='COMPLETED')
        
        # Use values().annotate() instead of Python-side loop
        summary_qs = transactions.values('payment_method').annotate(
            amount=Sum('total_amount'),
            count=Count('id')
        )
        
        return Response({
            'payment_methods': {
                entry['payment_method']: {'amount': str(entry['amount']), 'count': entry['count']}
                for entry in summary_qs
            }
        })

    @action(detail=False, methods=['get'], url_path='export_data')
    def export_data(self, request):
        """Export transaction data to CSV or PDF"""
        if not hasattr(request.user, 'business'):
            return Response({'error': 'Tidak ada bisnis yang terkait'}, status=status.HTTP_400_BAD_REQUEST)
            
        export_format = request.query_params.get('format', 'csv').lower()
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        queryset = self.get_queryset().filter(status='COMPLETED').order_by('-transaction_date')
        if start_date:
            queryset = queryset.filter(transaction_date__gte=start_date)
        if end_date:
            # include entire end_date
            queryset = queryset.filter(transaction_date__lte=f"{end_date} 23:59:59")
            
        if export_format in ['csv', 'excel', 'xlsx']:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laporan Penjualan"

            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid") # Emerald
            header_align = Alignment(horizontal="center", vertical="center")

            headers = ['No', 'Kode Transaksi', 'Tanggal & Waktu', 'Kasir', 'Metode Bayar', 'Total Pembayaran (Rp)', 'Status']
            ws.append(headers)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            total_amount_sum = 0
            for idx, txn in enumerate(queryset, start=1):
                amt = float(txn.total_amount)
                total_amount_sum += amt
                ws.append([
                    idx,
                    txn.transaction_code,
                    txn.transaction_date.strftime('%Y-%m-%d %H:%M:%S'),
                    txn.cashier_name,
                    txn.payment_method,
                    amt, # Numeric value for SUM formulas in Excel!
                    txn.status
                ])

            # Total Summary row
            ws.append(['', f'TOTAL ({len(queryset)} TRANSAKSI)', '', '', '', total_amount_sum, ''])

            # Style cells & format number
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                is_total = (row_idx == ws.max_row)
                for cell in row:
                    cell.border = thin_border
                    if is_total:
                        cell.font = Font(name="Calibri", size=11, bold=True, color="059669")
                        cell.fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
                    if cell.column == 6: # Total Pembayaran
                        cell.number_format = '#,##0'

            # Dynamic Auto-fit Column Width (wch + padding)!
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            filename = "transactions_export.xlsx" if export_format in ['excel', 'xlsx'] else "transactions_export.csv"
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
            
        elif export_format == 'pdf':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="transactions_export.pdf"'
            
            doc = SimpleDocTemplate(response, pagesize=landscape(A4))
            elements = []
            
            styles = getSampleStyleSheet()
            title = Paragraph(f"Laporan Penjualan - {request.user.business.name}", styles['Title'])
            elements.append(title)
            
            data = [['Kode Transaksi', 'Tanggal', 'Kasir', 'Metode Bayar', 'Total (Rp)']]
            
            for txn in queryset:
                data.append([
                    txn.transaction_code,
                    txn.transaction_date.strftime('%Y-%m-%d %H:%M'),
                    txn.cashier_name,
                    txn.payment_method,
                    f"{txn.total_amount:,.0f}"
                ])
                
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')), # emerald-600
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(table)
            doc.build(elements)
            return response
            
        else:
            return Response({'error': 'Format export tidak didukung'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def profit_loss(self, request):
        """
        Laporan Laba Rugi (Profit & Loss)
        Calculates Revenue, COGS, Gross Profit from cost_per_unit data.
        Supports date range filtering via ?start_date=&end_date=
        """
        if not hasattr(request.user, 'business'):
            return Response({'error': 'Tidak ada bisnis yang terkait'}, status=status.HTTP_400_BAD_REQUEST)

        business = request.user.business
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Base queryset: completed transactions for this business
        trx_qs = Transaction.objects.filter(
            business=business,
            status='COMPLETED'
        )

        if start_date:
            trx_qs = trx_qs.filter(transaction_date__date__gte=start_date)
        if end_date:
            trx_qs = trx_qs.filter(transaction_date__date__lte=end_date)

        # Overall aggregation
        overall = trx_qs.aggregate(
            total_revenue=Sum('total_amount'),
            total_discount=Sum('discount_amount'),
            transaction_count=Count('id')
        )

        total_revenue = overall['total_revenue'] or Decimal('0.00')
        total_discount = overall['total_discount'] or Decimal('0.00')
        transaction_count = overall['transaction_count'] or 0

        # COGS from TransactionItems
        items_qs = TransactionItem.objects.filter(
            transaction__in=trx_qs
        )

        cogs_data = items_qs.aggregate(
            total_cogs=Sum(F('cost_per_unit') * F('quantity')),
            total_items_sold=Sum('quantity')
        )

        total_cogs = cogs_data['total_cogs'] or Decimal('0.00')
        total_items_sold = cogs_data['total_items_sold'] or 0

        gross_profit = total_revenue - total_cogs
        margin_pct = (gross_profit / total_revenue * 100) if total_revenue > 0 else Decimal('0.00')

        # Per-product breakdown
        product_breakdown = items_qs.values(
            'product__id', 'product__name', 'product__code'
        ).annotate(
            qty_sold=Sum('quantity'),
            revenue=Sum('subtotal'),
            cogs=Sum(F('cost_per_unit') * F('quantity')),
        ).order_by('-revenue')

        products = []
        for p in product_breakdown:
            rev = p['revenue'] or Decimal('0.00')
            cogs = p['cogs'] or Decimal('0.00')
            profit = rev - cogs
            margin = (profit / rev * 100) if rev > 0 else Decimal('0.00')
            products.append({
                'product_id': p['product__id'],
                'product_name': p['product__name'],
                'product_code': p['product__code'],
                'qty_sold': p['qty_sold'],
                'revenue': str(rev),
                'cogs': str(cogs),
                'profit': str(profit),
                'margin_pct': round(float(margin), 1),
            })

        return Response({
            'period': {
                'start_date': start_date or 'all',
                'end_date': end_date or 'all',
            },
            'summary': {
                'total_revenue': str(total_revenue),
                'total_cogs': str(total_cogs),
                'gross_profit': str(gross_profit),
                'margin_pct': round(float(margin_pct), 1),
                'total_discount': str(total_discount),
                'transaction_count': transaction_count,
                'total_items_sold': total_items_sold,
            },
            'products': products,
        })
